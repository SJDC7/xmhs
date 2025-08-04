import random
from tkinter import *
from tkinter import messagebox, filedialog
from tkinter.ttk import *
from database import select_employee_project_costs
from data_import import import_data_from_excel_gs, import_data_from_excel


class WinGUI(Tk):
    def __init__(self):
        super().__init__()
        self.__win()
        self.tk_label_m373wtrj = self.__tk_label_m373wtrj(self)
        self.tk_label_m373y39h = self.__tk_label_m373y39h(self)
        self.tk_select_box_m373ypkv = self.__tk_select_box_m373ypkv(self)
        self.tk_select_box_m373z5oe = self.__tk_select_box_m373z5oe(self)
        self.tk_button_m374045p = self.__tk_button_m374045p(self)
        self.tk_button_m3740tmf = self.__tk_button_m3740tmf(self)
        self.tk_label_m37424uc = self.__tk_label_m37424uc(self)
        self.tk_select_box_m3743ynv = self.__tk_select_box_m3743ynv(self)
        self.tk_table_m3745jwl = self.__tk_table_m3745jwl(self)
        self.tk_button_m40uz2ao = self.__tk_button_m40uz2ao(self)

    def __win(self):
        self.title("农业灌区部门项目核算")
        # 设置窗口大小、居中
        width = 1060
        height = 650
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        geometry = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(geometry)

        self.resizable(width=False, height=False)

    def scrollbar_autohide(self, vbar, hbar, widget):
        """自动隐藏滚动条"""

        def show():
            if vbar: vbar.lift(widget)
            if hbar: hbar.lift(widget)

        def hide():
            if vbar: vbar.lower(widget)
            if hbar: hbar.lower(widget)

        hide()
        widget.bind("<Enter>", lambda e: show())
        if vbar: vbar.bind("<Enter>", lambda e: show())
        if vbar: vbar.bind("<Leave>", lambda e: hide())
        if hbar: hbar.bind("<Enter>", lambda e: show())
        if hbar: hbar.bind("<Leave>", lambda e: hide())
        widget.bind("<Leave>", lambda e: hide())

    def v_scrollbar(self, vbar, widget, x, y, w, h, pw, ph):
        widget.configure(yscrollcommand=vbar.set)
        vbar.config(command=widget.yview)
        vbar.place(relx=(w + x) / pw, rely=y / ph, relheight=h / ph, anchor='ne')

    def h_scrollbar(self, hbar, widget, x, y, w, h, pw, ph):
        widget.configure(xscrollcommand=hbar.set)
        hbar.config(command=widget.xview)
        hbar.place(relx=x / pw, rely=(y + h) / ph, relwidth=w / pw, anchor='sw')

    def create_bar(self, master, widget, is_vbar, is_hbar, x, y, w, h, pw, ph):
        vbar, hbar = None, None
        if is_vbar:
            vbar = Scrollbar(master)
            self.v_scrollbar(vbar, widget, x, y, w, h, pw, ph)
        if is_hbar:
            hbar = Scrollbar(master, orient="horizontal")
            self.h_scrollbar(hbar, widget, x, y, w, h, pw, ph)
        self.scrollbar_autohide(vbar, hbar, widget)

    def __tk_label_m373wtrj(self, parent):
        label = Label(parent, text="员工姓名：", anchor="center", )
        label.place(x=40, y=20, width=61, height=30)
        return label

    def __tk_label_m373y39h(self, parent):
        label = Label(parent, text="项目名称：", anchor="center", )
        label.place(x=265, y=19, width=62, height=30)
        return label

    def __tk_select_box_m373ypkv(self, parent):
        cb = Combobox(parent, state="readonly", )
        cb['values'] = (
            "全部", "王健壮", "梁皓", "徐康尊", "周靖", "孙雯", "李帆", "黄晓灿", "常雷", "孙碧伯", "孙慧杰",
            "李胤杭", "刘海岭", "张迪",  "郝志平","李晨")
        cb.place(x=104, y=20, width=150, height=30)
        return cb

    def __tk_select_box_m373z5oe(self, parent):
        cb = Combobox(parent, state="readonly", )
        cb['values'] = ("", "Python", "Tkinter Helper")
        cb.place(x=330, y=20, width=150, height=30)
        return cb

    def __tk_button_m374045p(self, parent):
        btn = Button(parent, text="查询", takefocus=False, command=self.query_data)
        btn.place(x=760, y=20, width=50, height=30)
        return btn

    def __tk_button_m3740tmf(self, parent):
        btn = Button(parent, text="数据导入", takefocus=False, command=self.open_import_window)
        btn.place(x=820, y=20, width=77, height=30)
        return btn

    def __tk_button_m40uz2ao(self, parent):
        btn = Button(parent, text="导出", takefocus=False, command=self.export_to_excel)
        btn.place(x=908, y=20, width=50, height=30)
        return btn

    def __tk_label_m37424uc(self, parent):
        label = Label(parent, text="日期选择：", anchor="center", )
        label.place(x=495, y=19, width=62, height=30)
        return label

    def __tk_select_box_m3743ynv(self, parent):
        cb = Combobox(parent, state="readonly", )
        cb['values'] = ("2024-11", "2024-12", "2025-01", "2025-02","2025-03","2025-04","2025-05","2025-06","2025-07")
        cb.place(x=560, y=19, width=150, height=30)
        return cb

    def __tk_table_m3745jwl(self, parent):
        # 表头字段 表头宽度
        columns = {"员工姓名": 90, "项目名称": 320,"产品名称": 120, "工时": 90, "人日": 90, "成本": 155, "年月": 103}
        tk_table = Treeview(parent, show="headings", columns=list(columns), )
        for text, width in columns.items():  # 批量设置列属性
            tk_table.heading(text, text=text, anchor='center')
            tk_table.column(text, anchor='center', width=width, stretch=True)  # stretch 不自动拉伸

        tk_table.place(x=21, y=81, width=1017, height=536)
        return tk_table

    def select_excel_file(self):
        file_path_cb = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path_cb:
            import_data_from_excel(file_path_cb)
            print("员工成本数据导入完成")

    def import_excel_file_gs(self):
        file_path_gs = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path_gs:
            import_data_from_excel_gs(file_path_gs)
            print("员工工时数据导入完成")

    def export_to_excel(self):
        """
        导出表格数据到 Excel 文件
        """
        # 弹出文件保存对话框
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not file_path:  # 如果用户取消保存
            return

        import openpyxl  # 确保已安装 openpyxl 库
        from openpyxl.styles import Alignment

        # 创建一个工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "导出数据"

        # 写入表头
        columns = self.tk_table_m3745jwl["columns"]
        for col_num, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 写入表格内容
        for row_num, item in enumerate(self.tk_table_m3745jwl.get_children(), start=2):
            row_values = self.tk_table_m3745jwl.item(item, "values")
            for col_num, value in enumerate(row_values, start=1):
                ws.cell(row=row_num, column=col_num, value=value)

        # 保存 Excel 文件
        try:
            wb.save(file_path)
            messagebox.showinfo("成功", f"数据已成功导出到:\n{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败:\n{str(e)}")

    def open_import_window(self):
        """
        弹出数据导入窗口
        :return:
        """
        import_window = Toplevel(self)
        import_window.title("数据导入")
        import_window.geometry("400x200")
        import_window.resizable(width=False, height=False)

        # 添加提示
        Label(import_window, text="请选择需要导入的时间和Excel文件", anchor="w").place(x=20, y=20, width=360, height=30)
        Button(import_window, text="成本导入", command=self.select_excel_file).place(x=140, y=60, width=100, height=30)
        Button(import_window, text="工时导入", command=self.import_excel_file_gs).place(x=140, y=100, width=100,
                                                                                        height=30)

    def query_data(self):
        """
        查询数据库并显示结果在列表中
        :return:
        """
        employee_name = self.tk_select_box_m373ypkv.get()
        timestamp = self.tk_select_box_m3743ynv.get()

        # 若未选员工姓名，弹出警告
        if not employee_name:
            messagebox.showwarning("警告", "请先选择员工姓名")
            return

        # 调用数据库查询函数
        results = select_employee_project_costs(employee_name, timestamp)

        # 清空表格旧数据
        for row in self.tk_table_m3745jwl.get_children():
            self.tk_table_m3745jwl.delete(row)

        total_hours = 0
        total_rr = 0
        total_cb = 0

        # 将查询结果插入表格中
        for results in results:
            total_hours += results[3]
            total_rr += results[4]
            total_cb += results[5]
            self.tk_table_m3745jwl.insert("", END, values=results)

        self.tk_table_m3745jwl.insert("", "end", values=("合计", "——", "——",total_hours, total_rr, total_cb, "——"))


class Win(WinGUI):
    def __init__(self, controller):
        self.ctl = controller
        super().__init__()
        self.__event_bind()
        self.__style_config()
        self.ctl.init(self)

    def __event_bind(self):
        pass

    def __style_config(self):
        pass


if __name__ == "__main__":
    win = WinGUI()
    win.mainloop()